import asyncio
import re
from io import BytesIO
from dataclasses import dataclass
from datetime import datetime

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlalchemy.ext.asyncio import AsyncSession

from bot.database.crud import cards as cards_crud
from bot.database.crud import character_arts as arts_crud
from bot.database.crud import characters as characters_crud
from bot.database.crud import contours as contours_crud
from bot.database.crud import trophies as trophies_crud
from bot.database.models import Card, CardOwnership, CardType, Character, CharacterArt, CharacterTrophy, Contour
from bot.services.errors import NotFoundError, ServiceError
from bot.services import art_storage_service, book_slot_service


@dataclass(frozen=True)
class SpreadsheetExport:
    filename: str
    data: bytes


async def export_character_cards(
    session: AsyncSession, character_id: int
) -> SpreadsheetExport:
    character = await characters_crud.get_by_id(session, character_id)
    if character is None:
        raise NotFoundError("Анкета не найдена.")
    ownerships = await cards_crud.list_character_ownerships(session, character_id)
    sheets = _character_card_sheets(character.name, ownerships)
    filename = (
        f"cards_{_safe_filename_part(character.name)}_{character.id}_"
        f"{datetime.now():%Y-%m-%d_%H-%M-%S}.xlsx"
    )
    return await _build(filename, sheets)


async def export_character_profile(
    session: AsyncSession, character_id: int
) -> SpreadsheetExport:
    character = await characters_crud.get_by_id(session, character_id)
    if character is None:
        raise NotFoundError("Анкета не найдена.")
    ownerships = await cards_crud.list_character_ownerships(session, character_id)
    contours = await contours_crud.list_for_character(session, character_id)
    arts = await arts_crud.list_for_character(session, character_id)
    trophies = await trophies_crud.list_for_character(session, character_id)
    slots = book_slot_service.calculate_usage(character, ownerships)
    sheets = [
        _profile_sheet(character, slots),
        _trophies_sheet(character, trophies),
        _arts_sheet(character, arts),
        _contours_sheet(character, contours),
        *_character_card_sheets(character.name, ownerships),
    ]
    filename = (
        f"character_{_safe_filename_part(character.name)}_{character.id}_"
        f"{datetime.now():%Y-%m-%d_%H-%M-%S}.xlsx"
    )
    return await _build(filename, sheets)


async def export_registry(session: AsyncSession) -> SpreadsheetExport:
    cards = await cards_crud.list_cards(
        session,
        limit=100_000,
        card_types=(CardType.SPECIAL, CardType.SPELL, CardType.CONTOUR, CardType.GM),
    )
    sheets = [
        _registry_sheet(cards, card_type, title)
        for card_type, title in (
            (CardType.SPECIAL, "Особые слоты"),
            (CardType.SPELL, "Заклинания"),
            (CardType.CONTOUR, "Контурные"),
            (CardType.GM, "ГеймМастерские"),
        )
    ]
    filename = f"cards_registry_{datetime.now():%Y-%m-%d_%H-%M-%S}.xlsx"
    return await _build(filename, sheets)


def _character_card_sheets(
    character_name: str, ownerships: list[CardOwnership]
) -> list[dict[str, object]]:
    locations = _ownership_locations(ownerships)
    return [
        _character_sheet(character_name, ownerships, card_type, title, locations)
        for card_type, title in (
            (CardType.SPECIAL, "Особые слоты"),
            (CardType.SPELL, "Заклинания"),
            (CardType.CONTOUR, "Контурные"),
            (CardType.ORDINARY, "Обычные"),
            (CardType.GM, "ГеймМастерские"),
        )
    ]


def _profile_sheet(character: Character, slots=None) -> dict[str, object]:
    special_used = getattr(slots, "special_used", 0)
    special_limit = getattr(slots, "special_limit", 100)
    free_used = getattr(slots, "free_used", 0)
    free_limit = getattr(slots, "free_limit", character.free_slot_limit or 10)
    rows = [
        ["Основное", "ID анкеты в БД", character.id],
        ["Основное", "Владелец VK", f"https://vk.ru/id{character.vk_id}"],
        ["Основное", "Имя персонажа", character.name],
        ["Основное", "Возраст", character.age if character.age is not None else ""],
        ["Основное", "Пол", character.gender],
        ["Основное", "Внешность", character.appearance],
        ["Основное", "Характер", character.personality],
        ["Основное", "Биография", character.biography],
        ["Статы", "Стрессоустойчивость", character.stress_resistance],
        ["Статы", "Речевой аппарат", character.speech],
        ["Статы", "Чуйка", character.intuition],
        ["Статы", "Хребет", character.spine],
        ["Статы", "Воля", character.will],
        ["Статы", "Нюх", character.scent],
        ["Навыки", "Навыки", character.skills],
        ["Прогресс", "Общий рейтинг", character.overall_rating.value],
        ["Прогресс", "Шакеи", character.shakei_balance],
        ["Прогресс", "Лимит Контуров", character.contour_limit],
        ["Книга", "Особые слоты", f"{special_used}/{special_limit}"],
        ["Книга", "Свободные слоты", f"{free_used}/{free_limit}"],
        ["Служебное", "Статус", "Подтверждена" if character.is_approved else "Не подтверждена"],
        ["Служебное", "Дата создания", _excel_datetime(character.created_at)],
        ["Дополнительно", "Дополнительно", character.additional],
    ]
    return {
        "name": "Анкета",
        "title": f"Полная анкета персонажа «{character.name}»",
        "subtitle": "Основные сведения, статы и прогресс персонажа",
        "headers": ["Раздел", "Поле", "Значение"],
        "rows": rows,
        "dateColumns": [2],
        "widths": [18, 25, 80],
    }


def _trophies_sheet(
    character: Character, trophies: list[CharacterTrophy]
) -> dict[str, object]:
    return {
        "name": "Трофеи",
        "title": f"Трофеи персонажа «{character.name}»",
        "subtitle": f"Всего постоянных трофеев: {len(trophies)}",
        "headers": ["ID трофея", "Ранг", "Название", "Описание", "Награда", "Выдал VK", "Дата"],
        "rows": [
            [
                trophy.id,
                trophy.rank.value,
                trophy.name,
                trophy.description,
                trophy.reward,
                f"https://vk.ru/id{trophy.awarded_by}",
                _excel_datetime(trophy.created_at),
            ]
            for trophy in trophies
        ],
        "emptyText": "Трофеев у персонажа пока нет",
        "dateColumns": [6],
        "widths": [12, 16, 28, 48, 36, 28, 20],
    }


def _contours_sheet(
    character: Character, contours: list[Contour]
) -> dict[str, object]:
    rows = []
    for contour in contours:
        if contour.components:
            composition = "\n".join(
                f"{component.position}. {component.ownership.display_name}"
                for component in contour.components
            )
        else:
            composition = contour.composition or "Требуется ручная привязка карт"
        rows.append(
            [
                contour.id,
                contour.slot,
                contour.name,
                f"{len(contour.components)}/{contour.card_capacity}",
                composition,
                contour.appearance,
                contour.primary_effect,
                contour.additional_capabilities,
                contour.activation_conditions,
                contour.duration,
                contour.conductivity,
                contour.overload_impact,
                _excel_datetime(contour.created_at),
            ]
        )
    return {
        "name": "Контуры",
        "title": f"Контуры персонажа «{character.name}»",
        "subtitle": f"Занято {len(contours)} из {character.contour_limit} доступных слотов",
        "headers": [
            "ID БД",
            "Слот",
            "Название",
            "Карт / размер",
            "Состав",
            "Внешний вид",
            "Основной эффект",
            "Дополнительные возможности",
            "Условия активации",
            "Продолжительность",
            "Проводимость",
            "Влияние на Перегрузку",
            "Дата создания",
        ],
        "rows": rows,
        "emptyText": "Контуров у персонажа пока нет",
        "dateColumns": [12],
        "widths": [10, 8, 24, 14, 32, 36, 42, 42, 36, 24, 24, 36, 20],
    }


def _arts_sheet(
    character: Character, arts: list[CharacterArt]
) -> dict[str, object]:
    rows = [
        [
            art.id,
            "Да" if art.is_primary else "Нет",
            art.caption,
            art.mime_type,
            art.width,
            art.height,
            art.file_size,
            art.sha256,
            art.created_by,
            _excel_datetime(art.created_at),
            "",
        ]
        for art in arts
    ]
    return {
        "name": "Арты",
        "title": f"Арты персонажа «{character.name}»",
        "subtitle": f"Всего изображений: {len(arts)}",
        "headers": [
            "ID арта",
            "Основной",
            "Подпись",
            "Формат",
            "Ширина",
            "Высота",
            "Размер, байт",
            "SHA-256",
            "Добавил VK ID",
            "Дата добавления",
            "Превью",
        ],
        "rows": rows,
        "dateColumns": [9],
        "widths": [10, 12, 36, 16, 12, 12, 16, 68, 18, 20, 28],
        "images": [
            {
                "row": index + 4,
                "column": 11,
                "data": art_storage_service.thumbnail_bytes(art.storage_key),
            }
            for index, art in enumerate(arts)
        ],
    }


def _character_sheet(
    character_name: str,
    ownerships: list[CardOwnership],
    card_type: CardType,
    title: str,
    locations: dict[int, str] | None = None,
) -> dict[str, object]:
    locations = locations or _ownership_locations(ownerships)
    selected = [item for item in ownerships if item.display_type is card_type]
    totals: dict[tuple[str, object], list[int]] = {}
    for item in selected:
        key = _ownership_group_key(item)
        counts = totals.setdefault(key, [0, 0, 0])
        counts[0] += 1
        if item.contour_component is None:
            counts[1] += 1
        else:
            counts[2] += 1
    headers = [
        "ID физической копии",
        "Игровой номер",
        "Название",
        "Вид / подтип",
        "Редкость",
        "Всего одинаковых",
        "Свободно",
        "Связано",
        "Описание",
        "Способ использования",
        "Дата получения",
        "Расположение в Книге",
    ]
    rows = []
    for ownership in selected:
        card = ownership.card
        game_number: int | str = ""
        if card is not None:
            game_number = (
                card.number
                if card.card_type is CardType.SPECIAL
                else card.registry_number if card.registry_number is not None else ""
            )
        status = locations[ownership.id]
        total, free, bound = totals[_ownership_group_key(ownership)]
        rows.append(
            [
                ownership.id,
                game_number,
                ownership.display_name,
                ownership.display_kind,
                ownership.display_rarity.value,
                total,
                free,
                bound,
                ownership.display_description,
                ownership.display_usage,
                _excel_datetime(ownership.obtained_at),
                status,
            ]
        )
    return {
        "name": title,
        "title": f"Карты персонажа «{character_name}» — {title}",
        "subtitle": f"Всего физических копий в категории: {len(rows)}",
        "headers": headers,
        "rows": rows,
        "dateColumns": [10],
        "widths": [18, 14, 24, 22, 10, 18, 12, 12, 42, 42, 20, 22],
    }


def _ownership_locations(ownerships: list[CardOwnership]) -> dict[int, str]:
    locations: dict[int, str] = {}
    first_special: set[int] = set()
    free: list[CardOwnership] = []
    for ownership in sorted(ownerships, key=lambda item: (item.obtained_at, item.id)):
        if ownership.contour_component is not None:
            locations[ownership.id] = f"Контур #{ownership.contour_component.contour_id}"
        elif (
            ownership.card is not None
            and ownership.card.card_type is CardType.SPECIAL
            and ownership.card_id not in first_special
        ):
            first_special.add(ownership.card_id)
            locations[ownership.id] = f"Особый слот #{ownership.card.number}"
        else:
            free.append(ownership)
    for index, ownership in enumerate(free, start=1):
        locations[ownership.id] = f"Свободный слот #{index}"
    return locations


def _ownership_group_key(ownership: CardOwnership) -> tuple[str, object]:
    if ownership.card_id is not None:
        return "registry", ownership.card_id
    return "ordinary", ownership.display_name.casefold()


def _registry_sheet(
    cards: list[Card], card_type: CardType, title: str
) -> dict[str, object]:
    selected = [card for card in cards if card.card_type is card_type]
    headers = [
        "Игровой номер",
        "ID БД",
        "Название",
        "Вид / подтип",
        "Редкость",
        "Описание",
        "Способ использования",
        "Лимит преобразований",
        "Живых копий",
        "Создана",
        "Создал VK",
    ]
    rows = [
        [
            card.number if card_type is CardType.SPECIAL else card.registry_number,
            card.id,
            card.name,
            card.kind,
            card.rarity.value,
            card.description,
            card.usage,
            card.transform_limit if card.transform_limit is not None else "Без лимита",
            card.copies_count,
            _excel_datetime(card.created_at),
            f"https://vk.ru/id{card.created_by}",
        ]
        for card in selected
    ]
    return {
        "name": title,
        "title": f"Реестр карт — {title}",
        "subtitle": f"Всего карт в категории: {len(rows)}",
        "headers": headers,
        "rows": rows,
        "dateColumns": [9],
        "widths": [14, 10, 24, 22, 10, 42, 42, 20, 14, 20, 28],
    }


async def _build(
    filename: str, sheets: list[dict[str, object]]
) -> SpreadsheetExport:
    try:
        return await asyncio.to_thread(_build_sync, filename, sheets)
    except (OSError, ValueError) as error:
        raise ServiceError(f"Не удалось создать XLSX: {error}") from error


def _build_sync(
    filename: str,
    sheets: list[dict[str, object]],
) -> SpreadsheetExport:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for index, definition in enumerate(sheets, start=1):
        _add_sheet(workbook, definition, index)
    stream = BytesIO()
    workbook.save(stream)
    return SpreadsheetExport(filename=filename, data=stream.getvalue())


def _add_sheet(workbook: Workbook, definition: dict[str, object], index: int) -> None:
    sheet = workbook.create_sheet(str(definition["name"]))
    headers = list(definition["headers"])
    rows = list(definition["rows"])
    widths = list(definition["widths"])
    display_widths = [min(float(width), 45) for width in widths]
    date_columns = set(definition["dateColumns"])
    last_column = get_column_letter(len(headers))

    sheet.merge_cells(f"A1:{last_column}1")
    sheet["A1"] = str(definition["title"])
    sheet["A1"].fill = PatternFill("solid", fgColor="6D174D")
    sheet["A1"].font = Font(name="Aptos Display", size=16, bold=True, color="FFFFFF")
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 30

    sheet.merge_cells(f"A2:{last_column}2")
    sheet["A2"] = str(definition.get("subtitle", ""))
    sheet["A2"].fill = PatternFill("solid", fgColor="F5E7F0")
    sheet["A2"].font = Font(name="Aptos", italic=True, color="4A1538")
    sheet["A2"].alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[2].height = 28

    header_fill = PatternFill("solid", fgColor="B52B7B")
    header_border = Border(bottom=Side(style="thin", color="7A1D57"))
    for column, value in enumerate(headers, start=1):
        cell = sheet.cell(row=3, column=column, value=value)
        cell.fill = header_fill
        cell.font = Font(name="Aptos", bold=True, color="FFFFFF")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = header_border
    sheet.row_dimensions[3].height = 30

    if rows:
        for row_index, values in enumerate(rows, start=4):
            for column_index, value in enumerate(values, start=1):
                cell = sheet.cell(row=row_index, column=column_index, value=value)
                cell.font = Font(name="Aptos", size=10)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if column_index - 1 in date_columns and isinstance(value, datetime):
                    cell.number_format = "dd.mm.yyyy hh:mm"
                if isinstance(value, str) and value.startswith("https://vk.ru/"):
                    cell.hyperlink = value
                    cell.style = "Hyperlink"
            sheet.row_dimensions[row_index].height = _content_row_height(
                values, display_widths
            )
        table = Table(
            displayName=f"CardsTable{index}",
            ref=f"A3:{last_column}{len(rows) + 3}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium4",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
    else:
        sheet.merge_cells(f"A4:{last_column}4")
        sheet["A4"] = str(
            definition.get("emptyText", "Карт в этой категории пока нет")
        )
        sheet["A4"].fill = PatternFill("solid", fgColor="FAF6F9")
        sheet["A4"].font = Font(name="Aptos", italic=True, color="6B5A65")
        sheet["A4"].alignment = Alignment(vertical="center")

    for column, width in enumerate(display_widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    for image_data in definition.get("images", []):
        image = ExcelImage(BytesIO(image_data["data"]))
        scale = min(160 / image.width, 160 / image.height, 1)
        image.width *= scale
        image.height *= scale
        anchor = f"{get_column_letter(int(image_data['column']))}{int(image_data['row'])}"
        sheet.add_image(image, anchor)
        sheet.row_dimensions[int(image_data["row"])].height = max(
            sheet.row_dimensions[int(image_data["row"])].height or 15,
            125,
        )
    sheet.freeze_panes = "A4"
    sheet.sheet_view.showGridLines = False
    if not rows:
        sheet.auto_filter.ref = f"A3:{last_column}3"
    sheet.print_title_rows = "1:3"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True


def _excel_datetime(value: datetime | None) -> datetime | str:
    if value is None:
        return ""
    return value.replace(tzinfo=None)


def _safe_filename_part(value: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", value.strip())
    cleaned = re.sub(r"\s+", "_", cleaned).strip("._")
    return cleaned[:64] or "character"


def _content_row_height(values: list[object], widths: list[object]) -> float:
    line_count = 1
    for value, width in zip(values, widths, strict=False):
        if value in (None, ""):
            continue
        text = str(value)
        column_width = max(float(width), 1)
        wrapped_lines = sum(
            max(1, (len(part) + int(column_width) - 1) // int(column_width))
            for part in text.splitlines() or [""]
        )
        line_count = max(line_count, wrapped_lines)
    return min(max(18, line_count * 15), 300)
