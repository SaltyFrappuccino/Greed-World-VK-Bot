from datetime import datetime
from io import BytesIO

from openpyxl import load_workbook

from bot.database.models import (
    Card,
    CardOwnership,
    CardType,
    Character,
    CharacterArt,
    Contour,
    ContourComponent,
    Rarity,
)
from bot.services.spreadsheet_service import (
    _build_sync,
    _arts_sheet,
    _character_sheet,
    _contours_sheet,
    _profile_sheet,
    _registry_sheet,
    _safe_filename_part,
)


def test_character_export_has_four_category_friendly_fields():
    ordinary = CardOwnership(
        card_id=None,
        character_id=1,
        ordinary_name="Верёвка",
        ordinary_kind="Инструмент",
        ordinary_description="Десять метров",
        ordinary_usage="Связать предметы",
        ordinary_rarity=Rarity.H,
        obtained_at=datetime(2026, 7, 18, 12, 30),
    )

    sheet = _character_sheet("Ава", [ordinary], CardType.ORDINARY, "Обычные")

    assert sheet["name"] == "Обычные"
    assert sheet["rows"][0][1] == ""
    assert sheet["rows"][0][2:5] == ["Верёвка", "Инструмент", "H"]
    assert sheet["rows"][0][5:8] == [1, 1, 0]
    assert sheet["dateColumns"] == [10]


def test_registry_export_uses_correct_game_number_and_vk_link():
    card = Card(
        id=41,
        name="Искра",
        card_type=CardType.SPELL,
        registry_number=0,
        kind="Заклинание",
        rarity=Rarity.H,
        description="Свет",
        usage="Назвать карту",
        copies_count=2,
        created_by=564059694,
        created_at=datetime(2026, 7, 18, 12, 30),
    )

    sheet = _registry_sheet([card], CardType.SPELL, "Заклинания")

    assert sheet["rows"][0][0:3] == [0, 41, "Искра"]
    assert sheet["rows"][0][-1] == "https://vk.ru/id564059694"


def test_full_profile_export_contains_all_sections_and_contour_components():
    character = Character(
        id=7,
        vk_id=564059694,
        name="Ава",
        age=28,
        gender="Женский",
        appearance="Описание внешности",
        personality="Описание характера",
        biography="Биография",
        skills="Навык",
        additional="Дополнение",
        stress_resistance=2,
        speech=3,
        intuition=4,
        spine=5,
        will=4,
        scent=3,
        overall_rating=Rarity.H,
        shakei_balance=100,
        contour_limit=3,
        is_approved=True,
        created_at=datetime(2026, 7, 18, 12, 30),
    )
    ownership = CardOwnership(
        ordinary_name="Верёвка",
        ordinary_kind="Инструмент",
        ordinary_rarity=Rarity.H,
    )
    contour = Contour(
        id=4,
        character_id=7,
        slot=1,
        card_capacity=2,
        name="Узел",
        created_by=564059694,
        created_at=datetime(2026, 7, 18, 13, 0),
    )
    contour.components = [ContourComponent(position=1, ownership=ownership)]

    profile = _profile_sheet(character)
    contours = _contours_sheet(character, [contour])

    assert ["Основное", "Владелец VK", "https://vk.ru/id564059694"] in profile["rows"]
    assert ["Статы", "Чуйка", 4] in profile["rows"]
    assert ["Дополнительно", "Дополнительно", "Дополнение"] in profile["rows"]
    assert contours["rows"][0][0:5] == [4, 1, "Узел", "1/2", "1. Верёвка"]

    export = _build_sync("character.xlsx", [profile, contours])
    workbook = load_workbook(BytesIO(export.data))
    profile_sheet = workbook["Анкета"]
    assert profile_sheet["C4"].value == 7
    assert profile_sheet["C4"].number_format == "General"
    assert profile_sheet["C25"].value == datetime(2026, 7, 18, 12, 30)
    assert profile_sheet["C25"].number_format == "dd.mm.yyyy hh:mm"


def test_character_filename_part_keeps_name_and_removes_forbidden_characters():
    assert _safe_filename_part("  Ава / Искра?  ") == "Ава___Искра"


def test_character_art_sheet_contains_metadata_and_preview(monkeypatch):
    buffer = BytesIO()
    from PIL import Image

    Image.new("RGB", (16, 12), "green").save(buffer, format="JPEG")
    monkeypatch.setattr(
        "bot.services.spreadsheet_service.art_storage_service.thumbnail_bytes",
        lambda _key: buffer.getvalue(),
    )
    character = Character(id=7, vk_id=100, name="Ава")
    art = CharacterArt(
        id=3,
        character_id=7,
        storage_key="characters/7/a.jpg",
        sha256="a" * 64,
        mime_type="image/jpeg",
        file_size=1234,
        width=1600,
        height=1200,
        caption="Портрет",
        is_primary=True,
        created_by=500,
        created_at=datetime(2026, 7, 19, 12, 0),
    )
    definition = _arts_sheet(character, [art])
    export = _build_sync("arts.xlsx", [definition])
    workbook = load_workbook(BytesIO(export.data))
    sheet = workbook["Арты"]

    assert sheet["A4"].value == 3
    assert sheet["B4"].value == "Да"
    assert sheet["C4"].value == "Портрет"
    assert len(sheet._images) == 1


def test_xlsx_is_built_with_openpyxl_tables_dates_and_freeze_panes():
    received_at = datetime(2026, 7, 18, 12, 30)
    definition = {
        "name": "Обычные",
        "title": "Карты персонажа «Ава» — Обычные",
        "subtitle": "Всего физических копий в категории: 1",
        "headers": ["Название", "Дата получения"],
        "rows": [["Верёвка", received_at]],
        "dateColumns": [1],
        "widths": [24, 20],
    }

    export = _build_sync("cards.xlsx", [definition])
    workbook = load_workbook(BytesIO(export.data))
    sheet = workbook["Обычные"]

    assert export.filename == "cards.xlsx"
    assert sheet["A1"].value == "Карты персонажа «Ава» — Обычные"
    assert sheet["B4"].value == received_at
    assert sheet["B4"].number_format == "dd.mm.yyyy hh:mm"
    assert sheet.freeze_panes == "A4"
    assert sheet.sheet_view.showGridLines is False
    assert list(sheet.tables) == ["CardsTable1"]
